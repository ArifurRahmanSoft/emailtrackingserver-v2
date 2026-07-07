"""Excel-backed storage for email open tracking."""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from threading import RLock
from typing import Final

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

TRACKING_COLUMNS: Final[tuple[str, ...]] = (
    "TrackingId",
    "OpenCount",
    "FirstOpen",
    "LastOpen",
    "LastIP",
    "UserAgent",
)


class TrackingStorageError(RuntimeError):
    """Raised when the tracking workbook cannot be read or written."""


@dataclass(frozen=True, slots=True)
class TrackingUpdateResult:
    """Result of one successful open-tracking update."""

    open_count: int
    status: str


class ExcelTrackingService:
    """Create and update the tracking workbook safely within one process."""

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self._lock = RLock()

    def initialize(self) -> None:
        """Create the workbook and normalize its columns when necessary."""
        with self._lock:
            try:
                self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
                if not self.workbook_path.exists():
                    workbook = Workbook()
                    worksheet = workbook.active
                    worksheet.title = "EmailTracking"
                    worksheet.append(TRACKING_COLUMNS)
                    workbook.save(self.workbook_path)
                    workbook.close()
                    return

                workbook = load_workbook(self.workbook_path)
                try:
                    worksheet = workbook.active
                    if self._ensure_columns(worksheet):
                        workbook.save(self.workbook_path)
                finally:
                    workbook.close()
            except (OSError, PermissionError, ValueError) as exc:
                raise TrackingStorageError(
                    f"Unable to initialize tracking workbook: {exc}"
                ) from exc

    def record_open(
        self,
        tracking_id: str,
        client_ip: str,
        user_agent: str,
        occurred_at: datetime | None = None,
    ) -> TrackingUpdateResult:
        """Insert or update a tracking row and save the workbook."""
        if not tracking_id.strip():
            raise ValueError("TrackingId must not be empty.")

        timestamp = occurred_at or datetime.now()
        with self._lock:
            try:
                self.initialize()
                workbook = load_workbook(self.workbook_path)
                try:
                    worksheet = workbook.active
                    matching_row = self._find_tracking_row(worksheet, tracking_id)
                    if matching_row is None:
                        open_count = 1
                        worksheet.append(
                            (
                                tracking_id,
                                open_count,
                                timestamp,
                                timestamp,
                                client_ip,
                                user_agent,
                            )
                        )
                        update_status = "created"
                    else:
                        count_cell = worksheet.cell(matching_row, 2)
                        open_count = self._normalize_count(count_cell.value) + 1
                        count_cell.value = open_count
                        worksheet.cell(matching_row, 4).value = timestamp
                        worksheet.cell(matching_row, 5).value = client_ip
                        worksheet.cell(matching_row, 6).value = user_agent
                        update_status = "updated"
                    workbook.save(self.workbook_path)
                finally:
                    workbook.close()
            except TrackingStorageError:
                raise
            except (OSError, PermissionError, ValueError) as exc:
                raise TrackingStorageError(
                    f"Unable to update tracking workbook: {exc}"
                ) from exc

        return TrackingUpdateResult(open_count=open_count, status=update_status)

    @staticmethod
    def _find_tracking_row(worksheet: Worksheet, tracking_id: str) -> int | None:
        """Return the row number for an exact TrackingId match."""
        for row_number in range(2, worksheet.max_row + 1):
            if worksheet.cell(row_number, 1).value == tracking_id:
                return row_number
        return None

    @staticmethod
    def _normalize_count(value: object) -> int:
        """Convert a stored count to a safe non-negative integer."""
        try:
            count = int(value) if value is not None else 0
        except (TypeError, ValueError):
            return 0
        return max(count, 0)

    @staticmethod
    def _ensure_columns(worksheet: Worksheet) -> bool:
        """Place required columns in order while preserving recognized data."""
        current_headers = tuple(
            worksheet.cell(1, column).value
            for column in range(1, worksheet.max_column + 1)
        )
        if current_headers == TRACKING_COLUMNS:
            return False

        header_positions = {
            str(header): index
            for index, header in enumerate(current_headers, start=1)
            if header is not None
        }
        preserved_rows = [
            [
                worksheet.cell(row, header_positions[column]).value
                if column in header_positions
                else None
                for column in TRACKING_COLUMNS
            ]
            for row in range(2, worksheet.max_row + 1)
        ]
        if worksheet.max_row:
            worksheet.delete_rows(1, worksheet.max_row)
        if worksheet.max_column > len(TRACKING_COLUMNS):
            worksheet.delete_cols(
                len(TRACKING_COLUMNS) + 1,
                worksheet.max_column - len(TRACKING_COLUMNS),
            )
        worksheet.append(TRACKING_COLUMNS)
        for row_values in preserved_rows:
            worksheet.append(row_values)
        return True
