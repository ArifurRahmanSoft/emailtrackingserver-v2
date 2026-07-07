"""Read-only access to tracking data for development diagnostics."""

from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

from app.services.excel_tracking import TRACKING_COLUMNS


class TrackingDebugService:
    """Read tracking records without creating or modifying the workbook."""

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path

    def read_records(self) -> list[dict[str, object]]:
        """Return all tracking rows using the required column names."""
        if not self.workbook_path.is_file():
            return []

        workbook = load_workbook(
            self.workbook_path,
            read_only=True,
            data_only=True,
        )
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            headers = next(rows, ())
            header_positions = {
                str(header): index
                for index, header in enumerate(headers)
                if header is not None
            }

            records: list[dict[str, object]] = []
            for row in rows:
                tracking_id_position = header_positions.get("TrackingId")
                tracking_id = (
                    row[tracking_id_position]
                    if tracking_id_position is not None
                    and tracking_id_position < len(row)
                    else None
                )
                if tracking_id in (None, ""):
                    continue

                record = {
                    column: self._serialize_value(
                        row[header_positions[column]]
                        if column in header_positions
                        and header_positions[column] < len(row)
                        else None
                    )
                    for column in TRACKING_COLUMNS
                }
                records.append(record)
            return records
        finally:
            workbook.close()

    def count_records(self) -> int:
        """Return the number of non-empty tracking records."""
        return len(self.read_records())

    @staticmethod
    def _serialize_value(value: object) -> object:
        """Convert Excel date/time values into stable JSON-friendly strings."""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, (date, time)):
            return value.isoformat()
        return value
