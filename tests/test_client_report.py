"""Tests for Version 2 client-based report APIs."""

from datetime import datetime, timedelta, timezone
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.api.client_report_routes as client_report_route_module
from app.models.campaign import Campaign, CampaignBase
from app.models.client_report import ClientReportResponse
from app.models.email_tracking import Base, EmailTracking
from app.services.client_report import (
    CLIENT_REPORT_EXPORT_CONTENT_TYPE,
    ClientReportService,
)
from main import app


def build_client_report_service() -> tuple[ClientReportService, sessionmaker]:
    """Create an isolated client-report service backed by shared in-memory SQLite."""
    engine = create_engine(
        "sqlite+pysqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    CampaignBase.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, expire_on_commit=False)
    service = ClientReportService(None)
    service._engine = engine
    service._session_factory = session_factory
    return service, session_factory


def seed_client_report_records(session_factory: sessionmaker) -> None:
    """Seed two clients with campaigns and tracking rows."""
    now = datetime(2026, 9, 1, 5, 0, tzinfo=timezone.utc)
    with session_factory() as session:
        session.add_all(
            [
                Campaign(
                    campaign_name="Power Campaign",
                    campaign_code="PC014",
                    client_code="USER001",
                ),
                Campaign(
                    campaign_name="CRM Campaign",
                    campaign_code="PC015",
                    client_code="USER001",
                ),
                Campaign(
                    campaign_name="Other Client",
                    campaign_code="PC999",
                    client_code="USER002",
                ),
            ]
        )
        session.add_all(
            [
                EmailTracking(
                    tracking_id="user001-open-reply",
                    campaign_code="PC014",
                    sender_mail="smtp-a@example.com",
                    project_name="PowerSoft",
                    open_count=2,
                    click_count=0,
                    download_count=0,
                    reply_count=1,
                    is_bounce=0,
                    created_at=now,
                ),
                EmailTracking(
                    tracking_id="user001-bounce-click-download",
                    campaign_code="PC014",
                    sender_mail="smtp-b@example.com",
                    project_name="PowerSoft",
                    open_count=0,
                    click_count=2,
                    download_count=1,
                    reply_count=0,
                    is_bounce=1,
                    created_at=now - timedelta(minutes=1),
                ),
                EmailTracking(
                    tracking_id="user001-second-campaign",
                    campaign_code="PC015",
                    sender_mail="smtp-a@example.com",
                    project_name="CRM",
                    open_count=1,
                    click_count=1,
                    download_count=0,
                    reply_count=0,
                    is_bounce=0,
                    created_at=now - timedelta(minutes=2),
                ),
                EmailTracking(
                    tracking_id="user002-private",
                    campaign_code="PC999",
                    sender_mail="smtp-a@example.com",
                    project_name="PowerSoft",
                    open_count=99,
                    click_count=99,
                    download_count=99,
                    reply_count=99,
                    is_bounce=1,
                    created_at=now,
                ),
            ]
        )
        session.commit()


def test_client_report_filters_by_client_code() -> None:
    service, session_factory = build_client_report_service()
    seed_client_report_records(session_factory)

    result = service.get_report(client_code="USER001")

    assert result.success is True
    assert result.client_code == "USER001"
    assert result.pagination.total_records == 3
    assert [row["tracking_id"] for row in result.data] == [
        "user001-open-reply",
        "user001-bounce-click-download",
        "user001-second-campaign",
    ]


def test_client_report_does_not_expose_other_client_campaigns() -> None:
    service, session_factory = build_client_report_service()
    seed_client_report_records(session_factory)

    result = service.get_report(client_code="USER001", campaign_code="PC999")

    assert result.pagination.total_records == 0
    assert result.data == []


def test_client_report_all_filters_work_with_and_condition() -> None:
    service, session_factory = build_client_report_service()
    seed_client_report_records(session_factory)

    result = service.get_report(
        client_code="USER001",
        sender_mail="smtp-b@example.com",
        campaign_code="PC014",
        project="PowerSoft",
        is_bounce="1",
        is_reply="false",
        is_open="false",
        is_click="true",
        is_download="true",
        from_date="2026-09-01",
        to_date="2026-09-01",
    )

    assert result.pagination.total_records == 1
    assert result.data[0]["tracking_id"] == "user001-bounce-click-download"


def test_client_report_default_pagination_returns_20_rows() -> None:
    service, session_factory = build_client_report_service()
    now = datetime(2026, 9, 1, 5, 0, tzinfo=timezone.utc)
    with session_factory() as session:
        session.add(Campaign(campaign_name="Bulk", campaign_code="PC014", client_code="USER001"))
        session.add_all(
            [
                EmailTracking(
                    tracking_id=f"bulk-{index:03d}",
                    campaign_code="PC014",
                    created_at=now - timedelta(minutes=index),
                )
                for index in range(25)
            ]
        )
        session.commit()

    result = service.get_report(client_code="USER001")

    assert result.pagination.page == 1
    assert result.pagination.per_page == 20
    assert result.pagination.total_records == 25
    assert result.pagination.total_pages == 2
    assert len(result.data) == 20


def test_client_report_custom_pagination_returns_total_count() -> None:
    service, session_factory = build_client_report_service()
    now = datetime(2026, 9, 1, 5, 0, tzinfo=timezone.utc)
    with session_factory() as session:
        session.add(Campaign(campaign_name="Bulk", campaign_code="PC014", client_code="USER001"))
        session.add_all(
            [
                EmailTracking(
                    tracking_id=f"bulk-{index:03d}",
                    campaign_code="PC014",
                    created_at=now - timedelta(minutes=index),
                )
                for index in range(25)
            ]
        )
        session.commit()

    result = service.get_report(client_code="USER001", page=2, per_page=10)

    assert result.pagination.page == 2
    assert result.pagination.per_page == 10
    assert result.pagination.total_records == 25
    assert result.pagination.total_pages == 3
    assert len(result.data) == 10
    assert result.data[0]["tracking_id"] == "bulk-010"


def test_client_report_no_campaigns_returns_empty_result() -> None:
    service, _ = build_client_report_service()

    result = service.get_report(client_code="USER001")

    assert result.success is True
    assert result.client_code == "USER001"
    assert result.pagination.total_records == 0
    assert result.pagination.total_pages == 0
    assert result.data == []


def test_client_report_endpoint_requires_client_code(
    monkeypatch,
) -> None:
    service, _ = build_client_report_service()
    monkeypatch.setattr(client_report_route_module, "client_report_service", service)
    client = TestClient(app)

    response = client.get("/api/client-report")

    assert response.status_code == 400
    assert response.json()["error"]["message"] == "client_code is required."


def test_client_report_endpoint_returns_response(monkeypatch) -> None:
    class FakeClientReportService:
        def get_report(
            self,
            client_code,
            page=1,
            per_page=20,
            sender_mail=None,
            campaign_code=None,
            project=None,
            is_bounce=None,
            is_reply=None,
            is_open=None,
            is_click=None,
            is_download=None,
            from_date=None,
            to_date=None,
        ):
            assert client_code == "USER001"
            assert sender_mail == "smtp-a@example.com"
            assert campaign_code == "PC014"
            return ClientReportResponse(
                success=True,
                client_code="USER001",
                pagination={
                    "page": page,
                    "per_page": per_page,
                    "total_records": 1,
                    "total_pages": 1,
                },
                data=[{"tracking_id": "row-1", "campaign_code": "PC014"}],
            )

    monkeypatch.setattr(
        client_report_route_module,
        "client_report_service",
        FakeClientReportService(),
    )
    client = TestClient(app)

    response = client.get(
        "/api/client-report",
        params={
            "client_code": "USER001",
            "sender_mail": "smtp-a@example.com",
            "campaign_code": "PC014",
        },
    )

    assert response.status_code == 200
    assert response.json()["data"][0]["tracking_id"] == "row-1"


def test_client_report_export_returns_xlsx_file() -> None:
    service, session_factory = build_client_report_service()
    seed_client_report_records(session_factory)

    result = service.export_report(client_code="USER001", campaign_code="PC014")

    workbook = load_workbook(BytesIO(result.content), read_only=True)
    worksheet = workbook["Client Report"]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[0])

    assert result.content_type == CLIENT_REPORT_EXPORT_CONTENT_TYPE
    assert result.row_count == 2
    assert "tracking_id" in headers
    assert [row[headers.index("tracking_id")] for row in rows[1:]] == [
        "user001-open-reply",
        "user001-bounce-click-download",
    ]


def test_client_report_export_endpoint_returns_attachment(monkeypatch) -> None:
    service, session_factory = build_client_report_service()
    seed_client_report_records(session_factory)
    monkeypatch.setattr(client_report_route_module, "client_report_service", service)
    client = TestClient(app)

    response = client.get("/api/client-report/export?client_code=USER001")

    assert response.status_code == 200
    assert response.headers["content-type"] == CLIENT_REPORT_EXPORT_CONTENT_TYPE
    assert response.headers["content-disposition"].startswith(
        'attachment; filename="Client_Report_'
    )
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert len(list(workbook["Client Report"].iter_rows(values_only=True))) == 4


def test_client_report_existing_campaign_rows_are_not_modified() -> None:
    service, session_factory = build_client_report_service()
    seed_client_report_records(session_factory)
    with session_factory() as session:
        before = list(
            session.execute(
                select(Campaign.campaign_code, Campaign.client_code).order_by(
                    Campaign.campaign_code
                )
            )
        )

    service.get_report(client_code="USER001")
    service.export_report(client_code="USER001")

    with session_factory() as session:
        after = list(
            session.execute(
                select(Campaign.campaign_code, Campaign.client_code).order_by(
                    Campaign.campaign_code
                )
            )
        )

    assert after == before
