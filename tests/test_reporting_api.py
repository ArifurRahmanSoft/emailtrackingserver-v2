"""Tests for the Version 2 reporting API."""

from datetime import datetime, timedelta, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

import app.api.routes as route_module
from app.models.email_tracking import Base, EmailTracking
from app.models.report import ReportFilterOptionsResponse, ReportResponse
from app.services.database_tracking import DatabaseTrackingService
from app.services.reporting import (
    REPORT_EXPORT_CONTENT_TYPE,
    ReportExportResult,
    ReportingService,
)
from main import app


BANGLADESH_TIMEZONE = ZoneInfo("Asia/Dhaka")


def build_reporting_service() -> tuple[ReportingService, sessionmaker, object]:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, expire_on_commit=False)

    database_service = DatabaseTrackingService(None)
    database_service._engine = engine
    database_service._session_factory = session_factory
    return ReportingService(database_service), session_factory, engine


def seed_records(
    session_factory: sessionmaker,
    count: int,
    start_time: datetime | None = None,
) -> None:
    base_time = start_time or datetime(2026, 7, 20, 10, 0, tzinfo=timezone.utc)
    with session_factory() as session:
        session.add_all(
            [
                EmailTracking(
                    tracking_id=f"report-{index:03d}",
                    sender_email=f"sender-{index}@example.com",
                    recipient_email=f"receiver-{index}@example.com",
                    project_name=f"Project {index % 3}",
                    created_at=base_time - timedelta(minutes=index),
                    open_count=index,
                    click_count=index % 5,
                    download_count=index % 4,
                    reply_count=index % 3,
                    is_bounce=1 if index % 7 == 0 else 0,
                )
                for index in range(count)
            ]
        )
        session.commit()


def seed_filter_records(session_factory: sessionmaker) -> None:
    base_time = datetime(2026, 7, 20, 10, 0, tzinfo=timezone.utc)
    rows = [
        EmailTracking(
            tracking_id="alpha-open-reply",
            sender_email="alpha@example.com",
            recipient_email="receiver-alpha@example.com",
            project_name="PowerSoft",
            created_at=base_time,
            open_count=5,
            click_count=0,
            download_count=0,
            reply_count=2,
            is_bounce=0,
        ),
        EmailTracking(
            tracking_id="alpha-click-download-bounce",
            sender_email="alpha@example.com",
            recipient_email="receiver-beta@example.com",
            project_name="PowerSoft",
            created_at=base_time - timedelta(minutes=1),
            open_count=1,
            click_count=3,
            download_count=4,
            reply_count=0,
            is_bounce=1,
        ),
        EmailTracking(
            tracking_id="beta-project",
            sender_email="beta@example.com",
            recipient_email="receiver-gamma@example.com",
            project_name="OtherProject",
            created_at=base_time - timedelta(minutes=2),
            open_count=0,
            click_count=0,
            download_count=0,
            reply_count=0,
            is_bounce=0,
        ),
        EmailTracking(
            tracking_id="gamma-click",
            sender_email="gamma@example.com",
            recipient_email="receiver-delta@example.com",
            project_name="PowerSoft",
            created_at=base_time - timedelta(minutes=3),
            open_count=0,
            click_count=1,
            download_count=0,
            reply_count=0,
            is_bounce=0,
        ),
    ]
    with session_factory() as session:
        session.add_all(rows)
        session.commit()


def export_rows(content: bytes) -> list[tuple[object, ...]]:
    """Return worksheet rows from an exported report workbook."""
    workbook = load_workbook(BytesIO(content), read_only=True)
    worksheet = workbook["Report"]
    return list(worksheet.iter_rows(values_only=True))


def test_default_pagination_returns_first_20_newest_rows() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 25)

    result = service.get_report()

    assert result.page == 1
    assert result.page_size == 20
    assert result.total_records == 25
    assert result.total_pages == 2
    assert result.has_next_page is True
    assert result.has_previous_page is False
    assert len(result.items) == 20
    assert [item.tracking_id for item in result.items[:3]] == [
        "report-000",
        "report-001",
        "report-002",
    ]


def test_custom_page_returns_next_rows() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 45)

    result = service.get_report(page=2)

    assert result.page == 2
    assert result.page_size == 20
    assert result.total_records == 45
    assert result.total_pages == 3
    assert result.has_next_page is True
    assert result.has_previous_page is True
    assert len(result.items) == 20
    assert result.items[0].tracking_id == "report-020"


def test_custom_page_size_returns_requested_number_of_rows() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 75)

    result = service.get_report(page_size=50)

    assert result.page == 1
    assert result.page_size == 50
    assert result.total_records == 75
    assert result.total_pages == 2
    assert len(result.items) == 50


def test_invalid_pagination_values_are_normalized() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 150)

    result = service.get_report(page=0, page_size=500)

    assert result.page == 1
    assert result.page_size == 100
    assert result.total_records == 150
    assert result.total_pages == 2
    assert len(result.items) == 100


def test_page_greater_than_total_pages_returns_empty_items() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 25)

    result = service.get_report(page=9)

    assert result.page == 9
    assert result.total_records == 25
    assert result.total_pages == 2
    assert result.has_next_page is False
    assert result.has_previous_page is True
    assert result.items == []


def test_empty_database_returns_empty_page() -> None:
    service, _, _ = build_reporting_service()

    result = service.get_report()

    assert result.page == 1
    assert result.page_size == 20
    assert result.total_records == 0
    assert result.total_pages == 0
    assert result.has_next_page is False
    assert result.has_previous_page is False
    assert result.items == []


def test_report_items_map_required_fields() -> None:
    service, session_factory, _ = build_reporting_service()
    created_at = datetime(2026, 7, 20, 10, 0, tzinfo=timezone.utc)
    with session_factory() as session:
        session.add(
            EmailTracking(
                tracking_id="field-map",
                sender_email="sender@example.com",
                recipient_email="receiver@example.com",
                project_name="Mapped Project",
                created_at=created_at,
                open_count=10,
                click_count=3,
                download_count=2,
                reply_count=1,
                is_bounce=1,
            )
        )
        session.commit()

    item = service.get_report().items[0]

    assert item.tracking_id == "field-map"
    assert item.sender_email == "sender@example.com"
    assert item.receiver_email == "receiver@example.com"
    assert item.project_name == "Mapped Project"
    assert item.send_date == created_at.astimezone(BANGLADESH_TIMEZONE)
    assert item.send_date.utcoffset() == timedelta(hours=6)
    assert item.open_count == 10
    assert item.click_count == 3
    assert item.download_count == 2
    assert item.reply_count == 1
    assert item.is_bounce is True

    with session_factory() as session:
        stored_created_at = (
            session.query(EmailTracking.created_at)
            .filter(EmailTracking.tracking_id == "field-map")
            .scalar()
        )
    assert stored_created_at.replace(tzinfo=timezone.utc) == created_at


def test_report_executes_one_count_and_one_select_query() -> None:
    service, session_factory, engine = build_reporting_service()
    seed_records(session_factory, 10)
    statements: list[str] = []

    @event.listens_for(engine, "before_cursor_execute")
    def capture_statement(conn, cursor, statement, parameters, context, executemany):
        statements.append(statement)

    service.get_report()
    event.remove(engine, "before_cursor_execute", capture_statement)

    assert len(statements) == 2
    assert statements[0].lstrip().upper().startswith("SELECT COUNT")
    assert "LIMIT" in statements[1].upper()


def test_no_filters_returns_all_records() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report()

    assert result.total_records == 4
    assert [item.tracking_id for item in result.items] == [
        "alpha-open-reply",
        "alpha-click-download-bounce",
        "beta-project",
        "gamma-click",
    ]


def test_sender_email_filter() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(sender_email="alpha@example.com")

    assert result.total_records == 2
    assert {item.tracking_id for item in result.items} == {
        "alpha-open-reply",
        "alpha-click-download-bounce",
    }


def test_project_name_filter() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(project_name="OtherProject")

    assert result.total_records == 1
    assert result.items[0].tracking_id == "beta-project"


def test_is_reply_filter() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(is_reply=True)

    assert result.total_records == 1
    assert result.items[0].tracking_id == "alpha-open-reply"


def test_is_bounce_filter() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(is_bounce=True)

    assert result.total_records == 1
    assert result.items[0].tracking_id == "alpha-click-download-bounce"


def test_is_open_filter() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(is_open=True)

    assert result.total_records == 2
    assert [item.tracking_id for item in result.items] == [
        "alpha-open-reply",
        "alpha-click-download-bounce",
    ]


def test_is_click_filter() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(is_click=True)

    assert result.total_records == 2
    assert [item.tracking_id for item in result.items] == [
        "alpha-click-download-bounce",
        "gamma-click",
    ]


def test_is_download_filter() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(is_download=True)

    assert result.total_records == 1
    assert result.items[0].tracking_id == "alpha-click-download-bounce"


def test_false_boolean_filters_are_ignored() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(
        is_reply=False,
        is_bounce=False,
        is_open=False,
        is_click=False,
        is_download=False,
    )

    assert result.total_records == 4


def test_empty_string_filters_are_ignored() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(sender_email="  ", project_name="")

    assert result.total_records == 4


def test_multiple_filters_work_together() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(
        sender_email="alpha@example.com",
        project_name="PowerSoft",
        is_open=True,
        is_download=True,
    )

    assert result.total_records == 1
    assert result.items[0].tracking_id == "alpha-click-download-bounce"


def test_no_matching_records_returns_empty_page() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.get_report(sender_email="missing@example.com")

    assert result.total_records == 0
    assert result.total_pages == 0
    assert result.items == []


def test_pagination_is_applied_after_filtering() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 30)

    result = service.get_report(page=2, page_size=5, project_name="Project 1")

    assert result.total_records == 10
    assert result.total_pages == 2
    assert len(result.items) == 5
    assert result.items[0].tracking_id == "report-016"


def test_report_endpoint_returns_paginated_response(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FakeReportingService:
        def get_report(
            self,
            page: int,
            page_size: int,
            sender_email: str | None = None,
            project_name: str | None = None,
            is_reply: bool = False,
            is_bounce: bool = False,
            is_open: bool = False,
            is_click: bool = False,
            is_download: bool = False,
        ) -> ReportResponse:
            assert page == 2
            assert page_size == 5
            assert sender_email == "sender@example.com"
            assert project_name is None
            assert is_reply is True
            assert is_bounce is False
            assert is_open is False
            assert is_click is False
            assert is_download is False
            return ReportResponse(
                page=2,
                page_size=5,
                total_records=6,
                total_pages=2,
                has_next_page=False,
                has_previous_page=True,
                items=[
                    {
                        "tracking_id": "endpoint-row",
                        "sender_email": "sender@example.com",
                        "receiver_email": "receiver@example.com",
                        "project_name": "Endpoint Project",
                        "send_date": datetime(
                            2026, 7, 20, 10, 0, tzinfo=timezone.utc
                        ),
                        "open_count": 1,
                        "click_count": 2,
                        "download_count": 3,
                        "reply_count": 4,
                        "is_bounce": False,
                    }
                ],
            )

    monkeypatch.setattr(route_module, "reporting_service", FakeReportingService())
    client = TestClient(app)

    response = client.get(
        "/api/report",
        params={
            "page": 2,
            "page_size": 5,
            "sender_email": "sender@example.com",
            "is_reply": "true",
        },
    )

    assert response.status_code == 200
    assert response.json()["page"] == 2
    assert response.json()["page_size"] == 5
    assert response.json()["total_records"] == 6
    assert response.json()["items"][0]["tracking_id"] == "endpoint-row"


def test_filter_options_normal_database_excludes_duplicates_and_blank_values() -> None:
    service, session_factory, _ = build_reporting_service()
    with session_factory() as session:
        session.add_all(
            [
                EmailTracking(
                    tracking_id="options-1",
                    sender_email="sales@company.com",
                    project_name="PowerSoft",
                ),
                EmailTracking(
                    tracking_id="options-2",
                    sender_email="marketing@company.com",
                    project_name="CRM",
                ),
                EmailTracking(
                    tracking_id="options-3",
                    sender_email="sales@company.com",
                    project_name="PowerSoft",
                ),
                EmailTracking(
                    tracking_id="options-4",
                    sender_email=" admin@company.com ",
                    project_name=" Email Automation ",
                ),
                EmailTracking(
                    tracking_id="options-5",
                    sender_email=None,
                    project_name=None,
                ),
                EmailTracking(
                    tracking_id="options-6",
                    sender_email="",
                    project_name="",
                ),
                EmailTracking(
                    tracking_id="options-7",
                    sender_email="   ",
                    project_name="   ",
                ),
            ]
        )
        session.commit()

    result = service.get_filter_options()

    assert result.sender_emails == [
        "admin@company.com",
        "marketing@company.com",
        "sales@company.com",
    ]
    assert result.project_names == [
        "CRM",
        "Email Automation",
        "PowerSoft",
    ]


def test_filter_options_empty_database_returns_empty_lists() -> None:
    service, _, _ = build_reporting_service()

    result = service.get_filter_options()

    assert result.sender_emails == []
    assert result.project_names == []


def test_filter_options_executes_distinct_queries_only() -> None:
    service, session_factory, engine = build_reporting_service()
    seed_filter_records(session_factory)
    statements: list[str] = []

    @event.listens_for(engine, "before_cursor_execute")
    def capture_statement(conn, cursor, statement, parameters, context, executemany):
        statements.append(statement)

    service.get_filter_options()
    event.remove(engine, "before_cursor_execute", capture_statement)

    assert len(statements) == 2
    assert all("DISTINCT" in statement.upper() for statement in statements)


def test_filter_options_endpoint_returns_dropdown_values(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FakeReportingService:
        def get_filter_options(self) -> ReportFilterOptionsResponse:
            return ReportFilterOptionsResponse(
                sender_emails=[
                    "admin@company.com",
                    "marketing@company.com",
                    "sales@company.com",
                ],
                project_names=["CRM", "Email Automation", "PowerSoft"],
            )

    monkeypatch.setattr(route_module, "reporting_service", FakeReportingService())
    client = TestClient(app)

    response = client.get("/api/report/filter-options")

    assert response.status_code == 200
    assert response.json() == {
        "sender_emails": [
            "admin@company.com",
            "marketing@company.com",
            "sales@company.com",
        ],
        "project_names": ["CRM", "Email Automation", "PowerSoft"],
    }


def test_export_all_records_returns_xlsx_with_all_database_columns() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 3)
    generated_at = datetime(2026, 7, 20, 10, 30, tzinfo=timezone.utc)

    result = service.export_report(generated_at=generated_at)
    rows = export_rows(result.content)
    headers = list(rows[0])
    tracking_id_index = headers.index("tracking_id")

    assert result.filename == "Report_20260720_103000.xlsx"
    assert result.content_type == REPORT_EXPORT_CONTENT_TYPE
    assert result.row_count == 3
    assert headers == DatabaseTrackingService.report_export_columns()
    assert [row[tracking_id_index] for row in rows[1:]] == [
        "report-000",
        "report-001",
        "report-002",
    ]


def test_export_filtered_by_sender_email() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.export_report(sender_email="alpha@example.com")
    rows = export_rows(result.content)
    tracking_id_index = list(rows[0]).index("tracking_id")

    assert result.row_count == 2
    assert [row[tracking_id_index] for row in rows[1:]] == [
        "alpha-open-reply",
        "alpha-click-download-bounce",
    ]


def test_export_filtered_by_project_name() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.export_report(project_name="OtherProject")
    rows = export_rows(result.content)
    tracking_id_index = list(rows[0]).index("tracking_id")

    assert result.row_count == 1
    assert rows[1][tracking_id_index] == "beta-project"


def test_export_filtered_by_is_reply() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.export_report(is_reply=True)
    rows = export_rows(result.content)
    tracking_id_index = list(rows[0]).index("tracking_id")

    assert result.row_count == 1
    assert rows[1][tracking_id_index] == "alpha-open-reply"


def test_export_filtered_by_is_bounce() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.export_report(is_bounce=True)
    rows = export_rows(result.content)
    tracking_id_index = list(rows[0]).index("tracking_id")

    assert result.row_count == 1
    assert rows[1][tracking_id_index] == "alpha-click-download-bounce"


def test_export_filtered_by_multiple_filters() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_filter_records(session_factory)

    result = service.export_report(
        sender_email="alpha@example.com",
        project_name="PowerSoft",
        is_open=True,
        is_download=True,
    )
    rows = export_rows(result.content)
    tracking_id_index = list(rows[0]).index("tracking_id")

    assert result.row_count == 1
    assert rows[1][tracking_id_index] == "alpha-click-download-bounce"


def test_export_empty_database_returns_header_only_workbook() -> None:
    service, _, _ = build_reporting_service()

    result = service.export_report()
    rows = export_rows(result.content)

    assert result.row_count == 0
    assert len(rows) == 1
    assert list(rows[0]) == DatabaseTrackingService.report_export_columns()


def test_export_large_database_exports_every_matching_record_without_pagination() -> None:
    service, session_factory, _ = build_reporting_service()
    seed_records(session_factory, 150)

    result = service.export_report()
    rows = export_rows(result.content)

    assert result.row_count == 150
    assert len(rows) == 151


def test_export_executes_single_filtered_select_query() -> None:
    service, session_factory, engine = build_reporting_service()
    seed_filter_records(session_factory)
    statements: list[str] = []

    @event.listens_for(engine, "before_cursor_execute")
    def capture_statement(conn, cursor, statement, parameters, context, executemany):
        statements.append(statement)

    service.export_report(sender_email="alpha@example.com", is_open=True)
    event.remove(engine, "before_cursor_execute", capture_statement)

    assert len(statements) == 1
    assert "WHERE" in statements[0].upper()
    assert "LIMIT" not in statements[0].upper()


def test_export_endpoint_returns_xlsx_attachment(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = load_workbook

    class FakeReportingService:
        def export_report(
            self,
            sender_email: str | None = None,
            project_name: str | None = None,
            is_reply: bool = False,
            is_bounce: bool = False,
            is_open: bool = False,
            is_click: bool = False,
            is_download: bool = False,
        ) -> ReportExportResult:
            assert sender_email == "alpha@example.com"
            assert project_name is None
            assert is_reply is True
            assert is_bounce is False
            assert is_open is False
            assert is_click is False
            assert is_download is False

            service, session_factory, _ = build_reporting_service()
            seed_filter_records(session_factory)
            return service.export_report(
                sender_email=sender_email,
                is_reply=is_reply,
                generated_at=datetime(2026, 7, 20, 10, 30, tzinfo=timezone.utc),
            )

    monkeypatch.setattr(route_module, "reporting_service", FakeReportingService())
    client = TestClient(app)

    response = client.get(
        "/api/report/export",
        params={"sender_email": "alpha@example.com", "is_reply": "true"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == REPORT_EXPORT_CONTENT_TYPE
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="Report_20260720_103000.xlsx"'
    )
    workbook(BytesIO(response.content), read_only=True)
